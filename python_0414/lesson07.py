# @time：2021/4/19 20:03
# @Aut：秋雨
# @File: .py
# @QQ：1957800403

'''
接口自动化测试步骤：
1、excel测试用例准备好，代码自动读取测试数据    ===OK
2、发送接口请求，得到响应数据       ===上上节课学习内容
3、断言：实际执行结果 vs 预期结果（用例文档中 ） ---通过 / 不通过  当前课程重点
4、最终结果，写入到Excel表格中 ---输出测试报告（目前课程无，高阶班才有）
'''

# 引入两个库
import requests  #第三方库
import openpyxl   #读取数据和写入数据的库

# ---读取测试用例
# 封装函数-测试用例的读取
def read_data(filename,sheetname):   #定义函数，参数
    wb = openpyxl.load_workbook(filename) #加载工作薄，打开一个已经存在excel文件
    sh = wb[sheetname]  #获取表单
    max_row = sh.max_row  #获取最大行数
    case_list = [ ]  #创建空列表，存储测试用例数据
    for i in range(2,max_row+1):
        dict1=dict(
        case_id = sh.cell(row=i,column=1).value,  #获取id
        url = sh.cell(row=i,column=5).value,   #获取url
        data = sh.cell(row=i,column=6).value,   #获取data
        expect = sh.cell(row=i,column=7).value  #获取期望
        )   #获取期望
        case_list.append(dict1)     #每循环一次，插入到list末尾
    return case_list      #定义返回值

#发送接口测试的函数
#封装函数
def api_fun(url,data):   #定义函数，参数
    # url_login = 'http://8.129.91.152:8766/futureloan/member/login'   #请求地址
    # data_login = {"mobile_phone": "15656541766","pwd": "lemon123"}   #请求正文
    headers_login ={'X-Lemonban-Media-Type':'lemonban.v2', 'Content-Type':'application/json'}  #请求头

    result = requests.post(url=url,json=data,headers=headers_login).json()   #调用post方法，返回值
    return result    #设置返回值

#写入excel测试结果
#写入的结果，进行封装
def wirte_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)  # 加载工作薄，打开一个已经存在excel文件
    sh = wb[sheetname]    #获取表单
    sh.cell(row=row,column=column).value = final_result   #直接对单元格的内容进行赋值/修改内容，写入结果
    wb.save(filename)   #保存文档

# # 接口测试实战
# # 第一步调用函数：读取数据read_data ，获取第一个表单re...，会返回测试用例的测试数据list，
#    # 定义变量cases接收返回值
# cases = read_data('test_case_api.xlsx','register')
# # print(cases)
# #每一条用例数据，拆分读取来 ，分别把url&data作为参数传给api_fun（）这个接口测试的函数
# for case in cases:         #依次访问cases中的数据 ，保存到定义的case变量中  #for循环
#     # print(case)           #打印结果  一整个字典数据
#     case_id = case['case_id']
#     url = case['url']         #做接口测试只需要获取它的url data就可以了，不需要整个字典数据
#     data = eval(case['data'])
#     # print(case_id,url,data)        #打印拿的url，data数据
#     # print(type(data))            #data数据串的类型
#
#     #获取期望结果，code msg信息
#     expect =eval(case['expect'])
#     expect_code = expect['code']
#     expect_msg = expect['msg']
#     print('预期结果code为：{}，msg为：{}'.format(expect_code,expect_msg))
#
#     #执行接口测试
#     # 获取case_id,url,data信息后，为了执行接口测试，获取执行接口测试   定义变量接收响应结果
#     real_result = api_fun(url=url,data=data)
#     # print(real_result)
#     #获取实际结果code、msg
#     real_code = real_result['code']
#     real_msg = real_result['msg']
#     print('实际结果code为：{}，msg为: {}'.format(real_code,real_msg))
#
#     # 断言：预期VS实际结果
#     if real_code == expect_code and real_msg == expect_msg:
#         print('这{}条测试用例执行通过！'.format(case_id))
#         final_re = 'passed'
#     else:
#         print('这{}条测试用例执行不通过！'.format(case_id))
#         final_re = 'Failed'
#     print('*'*50)
#
#     #写入最终的测试结果到excel测试中，文件名，表单名,用例id,结果列8,参数值
#     wirte_result('test_case_api.xlsx','register',case_id+1,8,final_re)
#
# # #补充：引号括起来字典是字符串格式
# # #eval()函数---运行被字符串包裹的表达式
# # # 例：'{"mobile_phone":"13652440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}'
# # str0 = '{"mobile_phone":"13652440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}'
# # dict0 = eval(str0)
# # print(type(dict0))
# # print(eval('2+3'))   #自行运算结果


# 接口测试实战,封装成函数
def execute_fun(filename,sheetname):    #文件名，表单名
    cases = read_data(filename,sheetname)
    # print(cases)
    for case in cases:         #依次访问cases中的数据 ，保存到定义的case变量中  #for循环
        # print(case)           #打印结果  一整个字典数据
        case_id = case['case_id']
        url = case['url']         #做接口测试只需要获取它的url data就可以了，不需要整个字典数据
        data = eval(case['data'])
        # print(case_id,url,data)        #打印拿的url，data数据
        # print(type(data))            #data数据串的类型

        #获取期望结果，code msg信息
        expect =eval(case['expect'])
        expect_code = expect['code']
        expect_msg = expect['msg']
        print('预期结果code为：{}，msg为：{}'.format(expect_code,expect_msg))

        #执行接口测试
        # 获取case_id,url,data信息后，为了执行接口测试，获取执行接口测试   定义变量接收响应结果
        real_result = api_fun(url=url,data=data)
        # print(real_result)
        #获取实际结果code、msg
        real_code = real_result['code']
        real_msg = real_result['msg']
        print('实际结果code为：{}，msg为: {}'.format(real_code,real_msg))

        # 断言：预期VS实际结果
        if real_code == expect_code and real_msg == expect_msg:
            print('这{}条测试用例执行通过！'.format(case_id))
            final_re = 'passed'
        else:
            print('这{}条测试用例执行不通过！'.format(case_id))
            final_re = 'Failed'
        print('*'*50)

        #写入最终的测试结果到excel测试中，文件名，表单名,用例id,结果列8,参数值
        wirte_result(filename,sheetname,case_id+1,8,final_re)
#调用接口自动化测试函数
# execute_fun('../test_data/test_case_api.xlsx', 'login')



